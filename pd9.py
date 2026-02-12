# 1. Dziennik użytkownika: Napisz program, który w pętli prosi użytkownika o wpisanie jednej
# linii tekstu. Każda wpisana linia powinna być dopisywana (tryb 'a' ) do pliku
# dziennik.txt . Program kończy działanie, gdy użytkownik wpisze "koniec".

while True:
    linia = input("Wpisz linię do dziennika (lub 'koniec' aby zakończyć): ")
    if linia.lower().strip() == "koniec":
        break
    with open('dziennik.txt', 'a', encoding='utf-8') as f:
        f.write(linia + '\n')

# 2. Licznik słów: Stwórz program, który pyta o nazwę pliku, odczytuje go, a następnie zlicza i
# wyświetla całkowitą liczbę słów w tym pliku. Obsłuż błąd FileNotFoundError , jeśli plik nie istnieje.

def word_counter():
    file_name = input("Podaj nazwę pliku: ")

    try:
        with open(file_name, "r", encoding="utf-8") as f:
            text = f.read()
            words = text.split()
            print(f"Liczba słów w pliku: {len(words)}")

    except FileNotFoundError:
        print("Błąd: Podany plik nie istnieje.")

word_counter()

# 3. Konfiguracja w JSON: Stwórz słownik Pythona z ustawieniami aplikacji, np.
# konfiguracja = {"uzytkownik": "admin", "motyw": "ciemny", "rozdzielczosc":
# [1920, 1080]} . Zapisz ten słownik do pliku config.json z wcięciami i poprawnym
# kodowaniem polskich znaków.

import json

configuration = {
    "uzytkownik": "admin",
    "motyw": "ciemny",
    "rozdzielczosc": [1920, 1080]
}

with open("config.json", "w", encoding="utf-8") as f:
    json.dump(configuration, f, ensure_ascii=False, indent=4)

print("Plik config.json został zapisany.")


# 4. Odczyt konfiguracji: Napisz program, który odczytuje plik config.json z poprzedniego
# zadania i wyświetla komunikat: Witaj, [uzytkownik]! Twój motyw to [motyw].

import json

try:
    with open("config.json", "r", encoding="utf-8") as f:
        configuration = json.load(f)

    user = configuration["uzytkownik"]
    motive= configuration["motyw"]

    print(f"Witaj, {user}! Twój motyw to {motive}.")

except FileNotFoundError:
    print("Błąd: Plik config.json nie istnieje.")

# 5. Eksport do CSV: Masz listę słowników: produkty = [{"nazwa": "Mleko", "cena":
# 3.50}, {"nazwa": "Chleb", "cena": 4.20}] . Zapisz te dane do pliku produkty.csv ,
# gdzie pierwszy wiersz to nagłówki ("nazwa", "cena").

import csv

produkty = [
    {"nazwa": "Mleko", "cena": 3.50},
    {"nazwa": "Chleb", "cena": 4.20}
]

with open("produkty.csv", "w", newline="", encoding="utf-8") as f:
    writer = csv.DictWriter(f, fieldnames=["nazwa", "cena"])
    writer.writeheader()
    writer.writerows(produkty)

print("Plik produkty.csv został zapisany.")

# 6. Import z CSV: Napisz program, który odczytuje plik produkty.csv i oblicza sumę cen
# wszystkich produktów. Użyj csv.DictReader , aby łatwiej odwoływać się do kolumn po nazwach.

import csv

suma = 0.0

try:
    with open("produkty.csv", "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)

        for wiersz in reader:
            suma += float(wiersz["cena"])

    print(f"Suma cen wszystkich produktów: {suma:.2f} zł")

except FileNotFoundError:
    print("Błąd: Plik produkty.csv nie istnieje.")


# 7. Tworzenie struktury folderów: Użyj modułu pathlib , aby napisać skrypt, który tworzy
# strukturę folderów: Projekt/src , Projekt/data , Projekt/docs .

from pathlib import Path


projekt = Path("Projekt")

(projekt / "src").mkdir(parents=True, exist_ok=True)
(projekt / "data").mkdir(parents=True, exist_ok=True)
(projekt / "docs").mkdir(parents=True, exist_ok=True)


# 8. Wyszukiwarka logów: Wyobraź sobie, że masz duży plik log.txt . Napisz program, który
# pyta użytkownika o szukane słowo (np. "ERROR") i zapisuje wszystkie linie zawierające to
# słowo do nowego pliku wyniki_wyszukiwania.txt .

lista = []

def wyszukiwarka_logow():
    szukane = input("Podaj szukane słowo (np. ERROR): ")

    with open("log.txt", "r", encoding="utf-8") as f:
        for linia in f:
            if szukane in linia:
                lista.append(linia)


wyszukiwarka_logow()

with open("wyniki_wyszukiwania.txt", "w", encoding="utf-8") as f:
    f.writelines(lista)

# 9. Prosty arkusz kalkulacyjny: Używając openpyxl , stwórz plik finanse.xlsx . W
# pierwszej kolumnie umieść nazwy wydatków (np. "Czynsz", "Jedzenie"), a w drugiej ich
# wartości. W komórce poniżej wartości oblicz i wstaw sumę wszystkich wydatków, używając
# formuły Excela (np. =SUM(B1:B2) ).

from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "Finanse"

wydatki = [
    ("Czynsz", 3000),
    ("Jedzenie", 1500)
]

for i, (nazwa, kwota) in enumerate(wydatki, start=1):
    ws[f"A{i}"] = nazwa
    ws[f"B{i}"] = kwota

wiersz_sumy = len(wydatki) + 1
ws[f"A{wiersz_sumy}"] = "Suma"
ws[f"B{wiersz_sumy}"] = f"=SUM(B1:B{len(wydatki)})"

wb.save("finanse.xlsx")

print("Plik finanse.xlsx został utworzony.")

# 10. Mini-projekt: Lista zadań: Stwórz prostą aplikację do zarządzania listą zadań. Program powinien:
# Przy starcie próbować wczytać zadania z pliku zadania.json .
# Pozwalać użytkownikowi dodać nowe zadanie.
# Pozwalać wyświetlić wszystkie zadania.
# Przy zamknięciu (lub na polecenie) zapisywać aktualną listę zadań do pliku zadania.json .

import json

file = "zadania.json"

def load_tasks():
    try:
        with open(file, "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        return []

def save_tasks(tasks):
    with open(file, "w", encoding="utf-8") as f:
        json.dump(tasks, f, ensure_ascii=False, indent=4)

def show_tasks(tasks):
    if not tasks:
        print("Brak zadań na liście.")
    else:
        print("\nLista zadań:")
        for i, task in enumerate(tasks, start=1):
            print(f"{i}. {task}")

def add_task(task):
    new_task = input("Wpisz treść nowego zadania: ")
    task.append(new_task)
    print("Zadanie dodane.")

def menu():
    print("""
1. Wyświetl zadania
2. Dodaj zadanie
3. Zapisz i zakończ
""")

def main():
    tasks = load_tasks()
    print("Witaj w liście zadań!")

    while True:
        menu()
        choice = input("Wybierz opcję: ")

        if choice == "1":
            show_tasks(tasks)
        elif choice == "2":
            add_task(tasks)
        elif choice == "3":
            save_tasks(tasks)
            print("Zadania zapisane. Do widzenia!")
            break
        else:
            print("Nieprawidłowy wybór.")

main()